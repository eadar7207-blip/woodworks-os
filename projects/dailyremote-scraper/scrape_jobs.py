import subprocess
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MAX_JOBS = 500
MAX_PAGES_PER_COUNTRY = 25
OUTPUT_FILE = "projects/dailyremote-scraper/europe_sales_jobs.xlsx"

EUROPEAN_COUNTRIES = [
    ("United Kingdom", "united-kingdom"),
    ("Germany", "germany"),
    ("France", "france"),
    ("Netherlands", "netherlands"),
    ("Spain", "spain"),
    ("Italy", "italy"),
    ("Poland", "poland"),
    ("Portugal", "portugal"),
    ("Sweden", "sweden"),
    ("Denmark", "denmark"),
    ("Ireland", "ireland"),
    ("Belgium", "belgium"),
    ("Norway", "norway"),
    ("Austria", "austria"),
    ("Switzerland", "switzerland"),
    ("Finland", "finland"),
    ("Romania", "romania"),
    ("Ukraine", "ukraine"),
    ("Greece", "greece"),
    ("Czech Republic", "czech-republic"),
    ("Hungary", "hungary"),
    ("Bulgaria", "bulgaria"),
    ("Croatia", "croatia"),
    ("Slovakia", "slovakia"),
    ("Serbia", "serbia"),
]

HEADERS = [
    "Job Title",
    "Country",
    "Employment Type",
    "Date Posted",
    "Location",
    "Salary",
    "Experience Level",
    "Categories",
    "Description Snippet",
    "Apply URL",
]

HEADER_BG = "1B3A6B"
HEADER_FG = "FFFFFF"


def scrape_page(country_slug, page_num):
    url = f"https://dailyremote.com/remote-sales-jobs-in-{country_slug}?page={page_num}"
    print(f"  [{country_slug}] page {page_num}: {url}")
    result = subprocess.run(
        ["firecrawl", "scrape", url],
        capture_output=True,
        text=True,
        timeout=90,
    )
    if result.returncode != 0:
        print(f"  ERROR: {result.stderr[:200]}")
        return []
    return parse_markdown(result.stdout)


def parse_markdown(markdown):
    jobs = []
    blocks = re.split(r'\n(?=## \[)', markdown)
    for block in blocks:
        if not block.startswith('## ['):
            continue
        if '[APPLY]' not in block:
            continue
        job = parse_job_block(block)
        if job and job['title']:
            jobs.append(job)
    return jobs


def parse_job_block(block):
    job = {
        'title': '',
        'employment_type': '',
        'date_posted': '',
        'location': '',
        'salary': '',
        'experience_level': '',
        'categories': '',
        'description_snippet': '',
        'job_url': '',
    }

    lines = block.strip().split('\n')

    title_match = re.match(r'^## \[(.+?)\]\((https?://[^\)]+)\)', lines[0])
    if not title_match:
        return None
    job['title'] = title_match.group(1).strip()
    job['job_url'] = title_match.group(2).strip()

    for line in lines[1:]:
        line = line.strip()
        if re.match(r'^(Full Time|Part Time|Internship|Contract)', line, re.IGNORECASE):
            job['employment_type'] = line
            break

    for line in lines:
        line = line.strip()
        if line.startswith('·'):
            job['date_posted'] = line.lstrip('·').strip()
            break

    for line in lines:
        if '⭐' not in line:
            continue

        loc_match = re.match(r'^(.*?)(?:💵|⭐)', line)
        if loc_match:
            loc = loc_match.group(1).strip()
            loc = re.sub(r'🌎\s*', '', loc).strip()
            job['location'] = loc

        salary_match = re.search(r'💵\s*([^⭐]+?)(?=⭐)', line)
        if salary_match:
            job['salary'] = salary_match.group(1).strip()

        exp_match = re.search(r'⭐\s*([^\[💼\n]+?)(?=\[|$)', line)
        if exp_match:
            job['experience_level'] = exp_match.group(1).strip()

        cats = re.findall(r'\[💼\s*([^\]]+)\]\([^\)]+\)', line)
        other_cats = re.findall(r'\[(?!💼)([^\]]+)\]\(https://dailyremote\.com/remote-[^\)]+\)', line)
        all_cats = cats + other_cats
        job['categories'] = ', '.join(all_cats) if all_cats else ''
        break

    desc_lines = []
    found_meta = False
    for line in lines:
        stripped = line.strip()
        if '⭐' in stripped:
            found_meta = True
            continue
        if not found_meta:
            continue
        if stripped.startswith('[APPLY]'):
            break
        if stripped and not stripped.startswith('[') and not stripped.startswith('#') and len(stripped) > 30:
            desc_lines.append(stripped)

    job['description_snippet'] = ' '.join(desc_lines[:3])
    return job


def build_excel(all_jobs, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Europe Sales Jobs"

    header_font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    url_font = Font(name="Calibri", color="4A9EDB", underline="single", size=10)
    row_font = Font(name="Calibri", size=10)
    wrap_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, job in enumerate(all_jobs, start=2):
        values = [
            job['title'],
            job.get('country', ''),
            job['employment_type'],
            job['date_posted'],
            job['location'],
            job['salary'],
            job['experience_level'],
            job['categories'],
            job['description_snippet'],
            job['job_url'],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap_align
            if col_idx == 10 and value:
                cell.hyperlink = value
                cell.font = url_font
            else:
                cell.font = row_font

    col_widths = [42, 18, 16, 14, 28, 24, 16, 32, 62, 52]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    wb.save(output_path)
    print(f"\nSaved: {output_path}")


def main():
    print(f"Starting Europe sales scrape — target: {MAX_JOBS} jobs\n")
    all_jobs = []
    seen_urls = set()

    for country_name, country_slug in EUROPEAN_COUNTRIES:
        if len(all_jobs) >= MAX_JOBS:
            break

        print(f"\n--- {country_name} ---")
        country_jobs_added = 0

        for page in range(1, MAX_PAGES_PER_COUNTRY + 1):
            if len(all_jobs) >= MAX_JOBS:
                break

            jobs = scrape_page(country_slug, page)

            if not jobs:
                print(f"  No jobs on page {page}, moving to next country")
                break

            new_jobs = 0
            for job in jobs:
                if len(all_jobs) >= MAX_JOBS:
                    break
                if job['job_url'] not in seen_urls:
                    seen_urls.add(job['job_url'])
                    job['country'] = country_name
                    all_jobs.append(job)
                    new_jobs += 1
                    country_jobs_added += 1

            print(f"  Page {page}: +{new_jobs} new jobs (total: {len(all_jobs)})")

            if new_jobs == 0:
                print(f"  No new unique jobs, moving to next country")
                break

        print(f"  {country_name} total: {country_jobs_added} jobs")

    print(f"\nTotal unique jobs scraped: {len(all_jobs)}")

    if not all_jobs:
        print("No jobs found.")
        sys.exit(1)

    build_excel(all_jobs, OUTPUT_FILE)
    print(f"Done. Open: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
