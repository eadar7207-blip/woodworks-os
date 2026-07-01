#!/usr/bin/env python3
"""
Collect 50 real estate broker leads via Firecrawl search.
Targets high-budget brokers (luxury, large teams, high volume).
Extracts phone/email from search snippets, then enriches via page scrape.

Usage:
  python3 scrape_brokers.py
"""

import subprocess
import re
import sys
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TARGET_LEADS = 50
OUTPUT_FILE = "projects/re-broker-leads/re_broker_leads.xlsx"

HEADER_BG = "1B3A6B"
HEADER_FG = "FFFFFF"
ACCENT_COLOR = "4A9EDB"

HEADERS = [
    "Full Name / Business",
    "Brokerage",
    "Phone Number",
    "Email",
    "City",
    "State",
    "Sales Volume / Indicator",
    "Team Size",
    "Specialty",
    "Budget Signal",
    "Website",
    "Source",
]

# Search queries designed to surface high-budget brokers with contact info
SEARCH_QUERIES = [
    # Luxury brokers with phones
    ("luxury real estate broker Chicago IL phone contact", "Chicago", "IL"),
    ("luxury real estate broker New York NY phone contact", "New York", "NY"),
    ("luxury real estate broker Los Angeles CA phone contact", "Los Angeles", "CA"),
    ("luxury real estate broker Miami FL phone contact", "Miami", "FL"),
    ("luxury real estate broker Dallas TX phone contact", "Dallas", "TX"),
    ("luxury real estate broker San Francisco CA phone contact", "San Francisco", "CA"),
    ("luxury real estate broker Boston MA phone contact", "Boston", "MA"),
    ("luxury real estate broker Seattle WA phone contact", "Seattle", "WA"),
    ("luxury real estate broker Atlanta GA phone contact", "Atlanta", "GA"),
    ("luxury real estate broker Houston TX phone contact", "Houston", "TX"),
    # Top producing broker teams
    ("top producing real estate team broker Chicago phone", "Chicago", "IL"),
    ("top producing real estate team broker Los Angeles phone", "Los Angeles", "CA"),
    ("top producing real estate team broker Miami phone", "Miami", "FL"),
    ("top producing real estate team broker New York phone", "New York", "NY"),
    ("top producing real estate team broker Dallas phone", "Dallas", "TX"),
    # High volume commercial/residential
    ("real estate brokerage team phone contact Las Vegas NV", "Las Vegas", "NV"),
    ("real estate brokerage team phone contact Phoenix AZ", "Phoenix", "AZ"),
    ("real estate brokerage team phone contact Denver CO", "Denver", "CO"),
    ("real estate brokerage team phone contact Nashville TN", "Nashville", "TN"),
    ("real estate brokerage team phone contact Austin TX", "Austin", "TX"),
    # Specific high-value niches
    ("Sotheby's real estate agent phone contact Chicago IL", "Chicago", "IL"),
    ("Compass real estate broker team phone New York NY", "New York", "NY"),
    ("Coldwell Banker luxury agent phone contact Los Angeles CA", "Los Angeles", "CA"),
    ("RE/MAX top agent phone contact Miami FL", "Miami", "FL"),
    ("Keller Williams top team phone contact Dallas TX", "Dallas", "TX"),
    # More cities
    ("luxury real estate broker Charlotte NC phone contact", "Charlotte", "NC"),
    ("luxury real estate broker San Diego CA phone contact", "San Diego", "CA"),
    ("luxury real estate broker Tampa FL phone contact", "Tampa", "FL"),
    ("luxury real estate broker Portland OR phone contact", "Portland", "OR"),
    ("luxury real estate broker Minneapolis MN phone contact", "Minneapolis", "MN"),
]

PHONE_RE = re.compile(r'(\+?1[\s\-\.]?)?\(?\d{3}\)?[\s\-\.]\d{3}[\s\-\.]\d{4}')
EMAIL_RE = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')
URL_RE = re.compile(r'URL:\s*(https?://[^\s]+)')
TITLE_RE = re.compile(r'^(.+?)\n', re.MULTILINE)


def run_search(query):
    """Run firecrawl search and return raw output."""
    result = subprocess.run(
        ["firecrawl", "search", query, "--limit", "5"],
        capture_output=True,
        text=True,
        timeout=60,
    )
    if result.returncode != 0:
        return ""
    return result.stdout


def parse_search_results(raw, city, state):
    """Parse firecrawl search output into lead dicts."""
    leads = []

    # Split into individual result blocks by blank lines + title pattern
    blocks = re.split(r'\n(?=[A-Z][^\n]+\n\s+URL:)', raw)

    for block in blocks:
        if not block.strip():
            continue
        lead = parse_result_block(block, city, state)
        if lead:
            leads.append(lead)

    return leads


def parse_result_block(block, city, state):
    """Extract lead data from a single search result block."""
    lines = block.strip().split('\n')
    if len(lines) < 2:
        return None

    snippet = block.lower()

    # Must be real estate related
    re_keywords = ['real estate', 'realty', 'broker', 'realtor', 'properties', 'homes', 'luxury']
    if not any(kw in snippet for kw in re_keywords):
        return None

    # Skip directories, aggregators, news articles
    skip = ['wikipedia', 'linkedin.com', 'facebook.com', 'instagram.com', 'twitter.com',
            'trulia', 'zillow.com', 'redfin', 'apartments.com', 'loopnet',
            'news', 'article', 'blog', 'magazine', 'rankings', 'best agents',
            'find an agent', 'search agents', 'agent finder']
    if any(s in snippet for s in skip):
        return None

    lead = {
        'name': '',
        'brokerage': '',
        'phone': '',
        'email': '',
        'city': city,
        'state': state,
        'sales_volume': '',
        'team_size': '',
        'specialty': '',
        'budget_signal': '',
        'website': '',
        'source': '',
    }

    # Title = first non-empty line
    for line in lines:
        line = line.strip()
        if line and not line.startswith('URL:') and not line.startswith('http'):
            lead['name'] = line[:80]
            break

    if not lead['name']:
        return None

    # URL
    url_match = URL_RE.search(block)
    if url_match:
        lead['website'] = url_match.group(1).strip()
        lead['source'] = url_match.group(1).strip()

    # Phone from snippet
    phone_match = PHONE_RE.search(block)
    if phone_match:
        lead['phone'] = phone_match.group(0).strip()

    # Email from snippet
    email_match = EMAIL_RE.search(block)
    if email_match:
        email = email_match.group(0)
        # Skip generic/noreply addresses
        if not any(bad in email.lower() for bad in ['noreply', 'no-reply', 'example', 'test@']):
            lead['email'] = email

    # Specialty
    if any(w in snippet for w in ['luxury', 'high-end', 'premium', 'prestige', 'estate']):
        lead['specialty'] = 'Luxury Residential'
    elif any(w in snippet for w in ['commercial', 'office', 'industrial', 'retail', 'investment']):
        lead['specialty'] = 'Commercial'
    elif any(w in snippet for w in ['team', 'group', 'associates', 'partners']):
        lead['specialty'] = 'Residential Team'
    else:
        lead['specialty'] = 'Residential'

    # Budget signals
    signals = []
    if 'luxury' in snippet:
        signals.append('luxury listings')
    if any(w in snippet for w in ['million', '$1', '$2', '$5', '$10']):
        signals.append('high-value transactions mentioned')
    if any(w in snippet for w in ['team', 'group', 'associates']):
        signals.append('team/group structure')
    if any(w in snippet for w in ['sotheby', 'compass', 'christie', 'berkshire']):
        signals.append('premium brokerage brand')
    if any(w in snippet for w in ['top producer', '#1', 'number one', 'award', 'ranked']):
        signals.append('top producer recognition')
    if any(w in snippet for w in ['zillow premier', 'featured agent', 'platinum']):
        signals.append('paid directory placement')

    lead['budget_signal'] = '; '.join(signals) if signals else 'luxury/high-volume search result'

    # Sales volume estimate from keywords
    if any(w in snippet for w in ['100 million', '$100m', 'billion']):
        lead['sales_volume'] = '$100M+'
    elif any(w in snippet for w in ['50 million', '$50m']):
        lead['sales_volume'] = '$50M+'
    elif any(w in snippet for w in ['20 million', '$20m']):
        lead['sales_volume'] = '$20M+'
    elif 'luxury' in snippet:
        lead['sales_volume'] = '$10M+ (luxury)'
    else:
        lead['sales_volume'] = 'High volume (est.)'

    # Team size
    team_match = re.search(r'(\d+)\s*(agent|associate|member|realtor)', snippet)
    if team_match:
        lead['team_size'] = team_match.group(1) + ' agents'
    elif any(w in snippet for w in ['team', 'group', 'associates']):
        lead['team_size'] = 'Team (est. 5+)'
    else:
        lead['team_size'] = 'Solo/Unknown'

    # Infer brokerage from name or snippet
    brand_map = {
        'sotheby': "Sotheby's International Realty",
        'compass': 'Compass',
        'keller williams': 'Keller Williams',
        're/max': 'RE/MAX',
        'remax': 'RE/MAX',
        'coldwell banker': 'Coldwell Banker',
        'century 21': 'Century 21',
        'berkshire': 'Berkshire Hathaway HomeServices',
        'exp realty': 'eXp Realty',
        'christie': "Christie's International",
        '@properties': '@properties',
    }
    brokerage = ''
    for key, brand in brand_map.items():
        if key in snippet:
            brokerage = brand
            break
    lead['brokerage'] = brokerage if brokerage else lead['name']

    return lead


def scrape_contact_page(url):
    """Try to scrape a single contact/about page for more info."""
    # Try contact page first
    contact_urls = [url, url.rstrip('/') + '/contact']
    for target_url in contact_urls:
        result = subprocess.run(
            ["firecrawl", "scrape", target_url],
            capture_output=True,
            text=True,
            timeout=60,
        )
        if result.returncode == 0 and result.stdout:
            content = result.stdout
            phone_match = PHONE_RE.search(content)
            email_match = EMAIL_RE.search(content)
            phone = phone_match.group(0).strip() if phone_match else ''
            email = email_match.group(0) if email_match else ''
            if phone or email:
                return phone, email
    return '', ''


def deduplicate(leads):
    """Deduplicate by phone, then by name."""
    seen_phones = set()
    seen_names = set()
    unique = []
    for lead in leads:
        phone_key = re.sub(r'\D', '', lead['phone']) if lead['phone'] else ''
        name_key = re.sub(r'\s+', ' ', lead['name'].lower().strip())

        if phone_key and len(phone_key) >= 10 and phone_key in seen_phones:
            continue
        if name_key in seen_names:
            continue

        if phone_key and len(phone_key) >= 10:
            seen_phones.add(phone_key)
        seen_names.add(name_key)
        unique.append(lead)
    return unique


def score_lead(lead):
    score = 0
    if lead['phone']:
        score += 15
    if lead['email']:
        score += 8
    if lead['website']:
        score += 3
    if 'Luxury' in lead['specialty']:
        score += 10
    if '$100M+' in lead['sales_volume']:
        score += 20
    elif '$50M+' in lead['sales_volume']:
        score += 12
    elif '$20M+' in lead['sales_volume']:
        score += 8
    if 'top producer' in lead['budget_signal'].lower() or 'ranked' in lead['budget_signal'].lower():
        score += 10
    if 'premium brokerage' in lead['budget_signal'].lower():
        score += 8
    return score


def build_excel(leads, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "RE Broker Leads"

    header_font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    url_font = Font(name="Calibri", color=ACCENT_COLOR, underline="single", size=10)
    row_font = Font(name="Calibri", size=10)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    alt_fill = PatternFill("solid", fgColor="EEF4FB")

    for row_idx, lead in enumerate(leads, start=2):
        values = [
            lead['name'],
            lead['brokerage'],
            lead['phone'],
            lead['email'],
            lead['city'],
            lead['state'],
            lead['sales_volume'],
            lead['team_size'],
            lead['specialty'],
            lead['budget_signal'],
            lead['website'],
            lead['source'],
        ]
        use_alt = (row_idx % 2 == 0)
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap_align
            if col_idx in (11, 12) and value and value.startswith('http'):
                cell.hyperlink = value
                cell.font = url_font
            else:
                cell.font = row_font
            if use_alt:
                cell.fill = alt_fill

    col_widths = [38, 32, 18, 30, 16, 8, 22, 14, 22, 48, 34, 34]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 24
    wb.save(output_path)
    print(f"\nSaved: {output_path}")


def main():
    print(f"Collecting real estate broker leads via Firecrawl search — target: {TARGET_LEADS}\n")
    all_leads = []

    for i, (query, city, state) in enumerate(SEARCH_QUERIES):
        if len(all_leads) >= TARGET_LEADS * 2:
            break

        print(f"[{i+1}/{len(SEARCH_QUERIES)}] {city}, {state}: {query[:60]}...")
        raw = run_search(query)
        leads = parse_search_results(raw, city, state)
        all_leads.extend(leads)
        print(f"  Got {len(leads)} leads (total: {len(all_leads)})")

        # Small delay to avoid rate limiting
        if i < len(SEARCH_QUERIES) - 1:
            time.sleep(1)

    print(f"\nRaw leads before dedup: {len(all_leads)}")

    unique_leads = deduplicate(all_leads)
    print(f"Unique leads after dedup: {len(unique_leads)}")

    # Enrich leads missing phone numbers by scraping their contact page
    missing_phone = [l for l in unique_leads if not l['phone'] and l['website']]
    print(f"\nEnriching {min(len(missing_phone), 20)} leads missing phone numbers...")
    enriched = 0
    for lead in missing_phone[:20]:
        phone, email = scrape_contact_page(lead['website'])
        if phone:
            lead['phone'] = phone
            enriched += 1
        if email and not lead['email']:
            lead['email'] = email
        time.sleep(0.5)

    print(f"  Enriched {enriched} leads with phone numbers")

    # Score and sort
    unique_leads.sort(key=score_lead, reverse=True)

    # Prioritize leads with phone numbers
    with_phone = [l for l in unique_leads if l['phone']]
    without_phone = [l for l in unique_leads if not l['phone']]
    final_leads = (with_phone + without_phone)[:TARGET_LEADS]

    print(f"\nFinal: {len(final_leads)} leads ({len([l for l in final_leads if l['phone']])} with phone numbers)")

    if not final_leads:
        print("No leads found.")
        sys.exit(1)

    build_excel(final_leads, OUTPUT_FILE)

    print(f"\n--- Top 15 Leads ---")
    for i, lead in enumerate(final_leads[:15], 1):
        phone_str = lead['phone'] or '(no phone)'
        print(f"{i:2}. {lead['name'][:35]:<35} | {phone_str:<16} | {lead['city']}, {lead['state']}")
    if len(final_leads) > 15:
        print(f"    ... and {len(final_leads) - 15} more")

    print(f"\nDone. Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
